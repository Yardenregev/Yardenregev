#import libraries 
from openpyxl import load_workbook
#Selecting specific sheet
#take input from user for file name
file_name_pull = input("Enter the file to read from: ")
file_name_pull += ".xlsx"
#load the workbook
wb = load_workbook(file_name_pull)
sheet = wb.active

cell_1 = sheet['E5'].value
cell_2 = sheet['G2'].value

#load another file and write cell_1 and cell_2 into it
file_name_push = input("Enter the file to write to: ")
file_name_push += ".xlsx"
wb2 = load_workbook(file_name_push)
sheet2 = wb2.active
sheet2['A1'].value = cell_1
sheet2['A2'].value = cell_2
wb2.save(file_name_push)


